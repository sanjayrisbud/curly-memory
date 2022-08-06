"""Defines the ReportsCreator class."""
from openpyxl import load_workbook

from charts.cash_vs_loan import CashVsLoanChart
from charts.summary_chart import SummaryChart


class ReportsCreator:
    """Class used to create the financial reports."""

    def __init__(self, statement, db_interface=None):
        self._statement = statement
        self._financial_data = None
        self._db_interface = db_interface
        self._charts_to_create = [
            ("SALN Chart", SummaryChart),
            ("Cash vs Loan Amount", CashVsLoanChart),
        ]

    @property
    def statement(self):
        """Return the financial statement."""
        return self._statement

    @property
    def financial_data(self):
        """Return the financial data."""
        if not self._financial_data:
            self.financial_data = self.get_previous_financial_data()
        return self._financial_data

    @financial_data.setter
    def financial_data(self, financial_data):
        """Set the financial data."""
        self._financial_data = financial_data

    def get_previous_financial_data(self):
        """Get the previous financial data."""
        if not self._db_interface:
            return [], {}, {}
        return self._db_interface.get_previous_financial_data()

    def run(self, data):
        """Perform class logic."""
        self.financial_data = data
        file_object = self._statement.file_object
        if not file_object.exists():
            workbook = load_workbook(
                self._statement.file_object.parent / "template.xlsx"
            )
        else:
            workbook = load_workbook(file_object)
        self.add_charts_to_workbook(workbook)
        workbook.save(file_object)

    def add_charts_to_workbook(self, workbook):
        """Add needed charts to the workbook."""
        for chart in self._charts_to_create:
            if chart[0] in workbook.get_sheet_names():
                sheet = workbook.get_sheet_by_name(chart[0])
                workbook.remove(sheet)
            sheet = workbook.create_sheet(title=chart[0])
            image = chart[1](self.financial_data, self._db_interface).get_image()
            image.anchor = "C1"
            sheet.add_image(image)
