"""Defines the StatementWriter class."""
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.borders import BORDER_NONE, BORDER_THIN
from openpyxl.utils import get_column_letter


class StatementWriter:
    """Class used to write out the financial _statement."""

    def __init__(self, statement):
        self._statement = statement
        self._financial_data = None
        self._row = 1

    @property
    def statement(self):
        """Return the financial statement."""
        return self._statement

    @property
    def financial_data(self):
        """Return the financial data."""
        return self._financial_data

    @financial_data.setter
    def financial_data(self, financial_data):
        """Set the financial data."""
        self._financial_data = financial_data

    def run(self, data):
        """Perform class logic."""
        self.financial_data = data
        workbook = load_workbook(self._statement.path / "template.xlsx")
        self.register_styles(workbook)
        assets = workbook.create_sheet("Assets")
        liabilities = workbook.create_sheet("Liabilities")

        self.populate_summary_sheet(workbook["Summary"])
        self.create_assets_sheet(assets)
        self._row = 1
        self.create_liabilities_sheet(liabilities)

        workbook.save(self._statement.file_object)

    def populate_summary_sheet(self, sheet):
        """Populate summary sheet with the mined _financial_data."""
        sheet["A2"] = "Sanjay Risbud"
        sheet["A3"] = self._statement.date.strftime("%B %Y")
        sheet["B26"] = self._statement.date.strftime("%m/%d/%Y")

        relevant_cells = {
            "BankAccounts": "B6",
            "Portfolio": "B8",
            "MutualFunds": "B9",
            "InsurancePolicies": "B11",
            "RealEstate": "B12",
            "CreditCard": "B16",
            "AmortizationSchedule": "B18",
        }

        for record in self._financial_data[0]:
            sheet[relevant_cells[record.title]] = record.value

    def create_assets_sheet(self, sheet):
        """Create assets sheet."""
        assets = self._financial_data[1]
        if not assets:
            return
        self.write_bank_accounts_info(sheet, assets.get("BankAccounts"))
        self.write_mutual_funds_and_bonds_info(sheet, assets.get("MutualFunds"))
        self.write_portfolio_info(sheet, assets.get("Portfolio"))
        self.write_life_insurance_info(sheet, assets.get("InsurancePolicies"))
        self.write_real_estate_info(sheet, assets.get("RealEstate"))
        self.adjust_column_widths(sheet)

    def create_liabilities_sheet(self, sheet):
        """Create liabilities sheet."""
        liabilities = self._financial_data[2]
        if not liabilities:
            return
        self.write_credit_card_info(sheet, liabilities["CreditCard"])
        self.write_current_loan_balance(sheet, liabilities["AmortizationSchedule"])
        self.adjust_column_widths(sheet)

    def write_bank_accounts_info(self, sheet, content):
        """Write out the information about bank accounts."""
        title = "Bank Accounts"
        header = ["Account Number", "Account Alias", "Bank", "Balance"]
        entries = [
            [record.account_number, record.account_alias, record.bank, record.balance]
            for record in content.get("records", [])
        ]
        self.create_section(
            sheet, title, header, entries, content.get("total_amount", "")
        )

    def write_mutual_funds_and_bonds_info(self, sheet, content):
        """Write out the information about mutual funds and bonds."""
        title = "Securities: bonds / mutual funds"
        header = ["Name of Security", "Company", "Market Value"]
        entries = [
            [record.stock, record.company, record.mkt_value]
            for record in content.get("records", [])
        ]
        self.create_section(
            sheet, title, header, entries, content.get("total_amount", "")
        )

    def write_portfolio_info(self, sheet, content):
        """Write out the information about stock portfolio."""
        title = "Stock in privately held companies"
        header = ["Name of Security", "Number of Shares", "Market Value"]
        entries = [
            [record.stock, record.shares, record.mkt_value]
            for record in content.get("records", [])
        ]
        self.create_section(
            sheet, title, header, entries, content.get("total_amount", "")
        )

    def write_life_insurance_info(self, sheet, content):
        """Write out the information about life insurance policies."""
        title = "Life Insurance"
        header = ["Policy Name", "Company", "Policy Number", "Face Amount"]
        entries = [
            [record.account_alias, record.bank, record.account_number, record.balance]
            for record in content.get("records", [])
        ]
        self.create_section(
            sheet, title, header, entries, content.get("total_amount", "")
        )

    def write_real_estate_info(self, sheet, content):
        """Write out the information about real estate."""
        title = "Real Estate"
        header = [
            "Description / Location",
            "Original Cost",
            "Purchase Date",
            "Market Value",
        ]
        entries = [
            [record.address, record.price, record.year_bought, record.market_value]
            for record in content.get("records", [])
        ]
        self.create_section(
            sheet, title, header, entries, content.get("total_amount", "")
        )

    def write_current_loan_balance(self, sheet, content):
        """Write out the information about current loan balance."""
        title = "Mortgage / real estate loans payable"
        header = [
            "Name of Creditor",
            "Original Amount",
            "Monthly Payment",
            "Amount Owing",
        ]
        entries = [
            ["Philippine Savings Bank", 1_960_000, record.amount, record.balance]
            for record in content.get("records", [])
        ]
        self.create_section(
            sheet, title, header, entries, content.get("total_amount", "")
        )

    def write_credit_card_info(self, sheet, content):
        """Write out the information about the credit card."""
        title = "Credit card & charge card debt"
        header = ["Account Number", "Bank", "Balance"]
        entries = [
            [record.account_number, record.bank, record.balance]
            for record in content.get("records", [])
        ]
        self.create_section(
            sheet, title, header, entries, content.get("total_amount", "")
        )

    def create_section(self, sheet, *args):
        """Create a section of the sheet."""
        title, header, entries, total = args
        sheet.cell(row=self._row, column=1, value=title).style = "title_style"

        self._row += 1
        for column, field in enumerate(header, start=1):
            sheet.cell(row=self._row, column=column, value=field).style = "header_style"

        i = 0
        for i, record in enumerate(entries, start=1):
            for column, field in enumerate(record, start=1):
                sheet.cell(
                    self._row + i, column=column, value=field
                ).style = "entry_style"

        self._row += i + 1
        sheet.cell(row=self._row, column=1, value="Total").style = "total_title_style"
        for i in range(2, len(header)):
            sheet.cell(row=self._row, column=i).style = "total_blank_cell"
        sheet.cell(row=self._row, column=i + 1, value=total).style = "total_value_style"

        self._row += 2

    @staticmethod
    def register_styles(workbook):
        """Register the styles to be used in the workbook."""
        dark_side = Side(border_style=BORDER_THIN, color="00000000")
        no_side = Side(border_style=BORDER_NONE)
        full_border = Border(
            left=dark_side, right=dark_side, top=dark_side, bottom=dark_side
        )
        no_border = Border(left=no_side, right=no_side, top=no_side, bottom=no_side)
        light_fill = PatternFill("solid", fgColor="00FFFFCC")
        olive_fill = PatternFill("solid", fgColor="00333300")

        blank_cell = NamedStyle("blank_cell")
        title_style = NamedStyle("title_style")
        header_style = NamedStyle("header_style")
        entry_style = NamedStyle("entry_style")
        total_title_style = NamedStyle("total_title_style")
        total_blank_cell = NamedStyle("total_blank_cell")
        total_value_style = NamedStyle("total_value_style")

        title_style.font = Font(bold=True, italic=True, size=12)
        header_style.font = Font(bold=True, size=10)
        entry_style.font = Font(size=10)
        total_title_style.font = Font(size=10, bold=True, color="FFFFFF")
        header_style.alignment = Alignment(horizontal="center")
        total_title_style.fill = total_blank_cell.fill = olive_fill
        header_style.fill = total_value_style.fill = light_fill
        title_style.border = blank_cell.border = no_border
        header_style.border = (
            entry_style.border
        ) = (
            total_title_style.border
        ) = total_blank_cell.border = total_value_style.border = full_border

        workbook.add_named_style(title_style)
        workbook.add_named_style(header_style)
        workbook.add_named_style(entry_style)
        workbook.add_named_style(total_title_style)
        workbook.add_named_style(total_value_style)
        workbook.add_named_style(blank_cell)
        workbook.add_named_style(total_blank_cell)

    @staticmethod
    def adjust_column_widths(sheet):
        """Adjust the sheet's column widths for sheet readability."""
        column_widths = []
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[get_column_letter(i + 1)].width = column_width
