"""Defines the StatementWriter class."""
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.borders import BORDER_NONE, BORDER_THIN
from openpyxl.utils import get_column_letter

from misc.file_processor import FileProcessor


class StatementWriter(FileProcessor):
    """Class used to write out the financial statement."""

    def __init__(self, filename, path, date):
        super().__init__(filename, path, date)
        self.data = None
        self.row = self.column = 1

    def run(self, data):
        """Perform class logic."""
        self.data = data
        template = load_workbook(self.file_object.parent / "template.xlsx")
        self.register_styles(template)
        assets = template.create_sheet("Assets")
        liabilities = template.create_sheet("Liabilities")

        self.populate_summary_sheet(template["Summary"])
        self.create_assets_sheet(assets)
        self.row = self.column = 1
        self.create_liabilities_sheet(liabilities)

        template.save(self.file_object)
        return self.archive()

    def populate_summary_sheet(self, sheet):
        """Populate summary sheet with the mined data."""
        sheet["A2"] = "Sanjay Risbud"
        sheet["A3"] = self.date.strftime("%B %Y")
        sheet["B26"] = self.date.strftime("%m/%d/%Y")

        relevant_cells = {
            "BankAccounts": "B6",
            "Portfolio": "B8",
            "MutualFunds": "B9",
            "InsurancePolicies": "B11",
            "RealEstate": "B12",
            "AmortizationSchedule": "B18",
        }
        for asset_class in self.data.get("asset_parsers", []):
            type_ = asset_class.__class__.__name__.replace("Parser", "")
            sheet[relevant_cells[type_]] = asset_class.total_amount
        for liability_class in self.data.get("liability_parsers", []):
            type_ = liability_class.__class__.__name__.replace("Parser", "")
            sheet[relevant_cells[type_]] = liability_class.total_amount

        if self.data.get("credit_card", None):
            sheet["B16"] = self.data["credit_card"].balance

    def create_assets_sheet(self, sheet):
        """Create assets sheet with the mined data."""
        self.write_mutual_funds_and_bonds_info(sheet, self.data["asset_parsers"][3])
        self.write_portfolio_info(sheet, self.data["asset_parsers"][1])
        self.adjust_column_widths(sheet)

    def create_liabilities_sheet(self, sheet):
        """Create liabilities sheet with the mined data."""

    def write_mutual_funds_and_bonds_info(self, sheet, *args):
        """Write out the information about mutual funds and bonds."""
        title = "Securities: bonds / mutual funds"
        header = ["Name of Security", "Company", "Market Value"]
        entries = [
            [record.stock, "Sunlife", record.mkt_value]
            for record in args[0].parsed_data
        ]
        self.create_section(sheet, title, header, entries, args[0].total_amount)

    def write_portfolio_info(self, sheet, *args):
        """Write out the information about stock portfolio."""
        title = "Stock in privately held companies"
        header = ["Name of Security", "Number of Shares", "Market Value"]
        entries = [
            [record.stock, record.shares, record.mkt_value]
            for record in args[0].parsed_data
        ]
        self.create_section(sheet, title, header, entries, args[0].total_amount)

    def create_section(self, sheet, *args):
        """Create a section of the sheet."""
        title, header, entries, total = args
        sheet.cell(row=self.row, column=1, value=title).style = "title_style"

        self.row += 1
        for column, field in enumerate(header, start=1):
            sheet.cell(row=self.row, column=column, value=field).style = "header_style"

        for i, record in enumerate(entries, start=1):
            for column, field in enumerate(record, start=1):
                sheet.cell(
                    self.row + i, column=column, value=field
                ).style = "entry_style"

        self.row += i + 1
        sheet.cell(row=self.row, column=1, value="Total").style = "total_title_style"
        for i in range(2, len(header)):
            sheet.cell(row=self.row, column=i).style = "total_blank_cell"
        sheet.cell(row=self.row, column=i + 1, value=total).style = "total_value_style"

        self.row += 2

    @staticmethod
    def register_styles(workbook):
        """Register the styles to be used in the workbook."""
        dark_side = Side(border_style=BORDER_THIN, color="00000000")
        no_side = Side(border_style=BORDER_NONE)
        full_border = Border(
            left=dark_side, right=dark_side, top=dark_side, bottom=dark_side
        )
        no_border = Border(left=no_side, right=no_side, top=no_side, bottom=no_side)
        light_fill = PatternFill("solid", fgColor="00FFFFCC")
        olive_fill = PatternFill("solid", fgColor="00333300")

        blank_cell = NamedStyle("blank_cell")
        title_style = NamedStyle("title_style")
        header_style = NamedStyle("header_style")
        entry_style = NamedStyle("entry_style")
        total_title_style = NamedStyle("total_title_style")
        total_blank_cell = NamedStyle("total_blank_cell")
        total_value_style = NamedStyle("total_value_style")

        title_style.font = Font(bold=True, italic=True, size=12)
        header_style.font = Font(bold=True, size=10)
        entry_style.font = Font(size=10)
        total_title_style.font = Font(size=10, bold=True, color="FFFFFF")
        header_style.alignment = Alignment(horizontal="center")
        total_title_style.fill = total_blank_cell.fill = olive_fill
        header_style.fill = total_value_style.fill = light_fill
        title_style.border = blank_cell.border = no_border
        header_style.border = (
            entry_style.border
        ) = (
            total_title_style.border
        ) = total_blank_cell.border = total_value_style.border = full_border

        workbook.add_named_style(title_style)
        workbook.add_named_style(header_style)
        workbook.add_named_style(entry_style)
        workbook.add_named_style(total_title_style)
        workbook.add_named_style(total_value_style)
        workbook.add_named_style(blank_cell)
        workbook.add_named_style(total_blank_cell)

    @staticmethod
    def adjust_column_widths(sheet):
        """Adjust the sheet's column widths for sheet readability."""
        column_widths = []
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[get_column_letter(i + 1)].width = column_width
